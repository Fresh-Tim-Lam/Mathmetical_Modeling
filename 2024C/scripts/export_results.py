# -*- coding: utf-8 -*-
"""结果导出：把求解方案填入附件3 模板（result1_1 / result1_2 / result2.xlsx）。

- 模板行布局：第1行表头（地块名 + 作物1..41）；行2~55 第一季 A1..F4；行56~83 第二季 D1..F4；
- 填写口径：作物 j 在对应(地块,季次)行的第 j+2 列填种植亩数（α×单元面积）；
- 单季露天（A/B/C）只填第一季；水浇地/大棚两季各填对应季次。
"""
import sys
from collections import defaultdict

import openpyxl
import pandas as pd

sys.path.insert(0, r'd:\AAA_Jupyter\BBB_Competition\2025\C\src')

from de_solver_full import FullDE, YEARS
from revenue import DISTS
from scheme import SchemeModel

BASE = r'd:\AAA_Jupyter\BBB_Competition\2025\C'
DOCS = f'{BASE}/docs'
OUT = f'{BASE}/data'


def build_baseline():
    sm = SchemeModel()
    plant = pd.read_csv(f'{BASE}/data/2023种植情况.csv', skipinitialspace=True)
    base = defaultdict(lambda: defaultdict(list))
    for _, r in plant.iterrows():
        plot = str(r['种植地块']).strip()
        s = 1 if str(r['种植季次']).strip() in ('单季', '第一季') else 2
        u = sm.unit_of(plot, s)
        alpha = float(r['种植面积/亩']) / sm.unit_area[u]
        base[plot][s].append((int(r['作物编号']), alpha))
    return sm, dict(base)


def row_map(ws):
    """模板行布局：{(季次, 地块): 行号}。"""
    m, season = {}, None
    for i in range(1, ws.max_row + 1):
        a = ws.cell(i, 1).value
        b = ws.cell(i, 2).value
        if a is not None:
            season = str(a).replace('\n', '')
        if season and b is not None:
            m[(season, str(b).strip())] = i
    return m


def fill_workbook(template, out_path, plan):
    """plan = {年份: YearPlan} → 填满模板所有 sheet 并保存。"""
    wb = openpyxl.load_workbook(template)
    for y in YEARS:
        ws = wb[str(y)]
        rm = row_map(ws)
        for p, seasons in plan[y].items():
            for s, crops in seasons.items():
                key = '第一季' if s == 1 else '第二季'
                if key == '第一季' and p[0] in 'ABC':
                    pass                                # A/B/C 单季只在此行
                row = rm.get((key, p))
                if row is None:
                    continue
                for j, alpha in crops:
                    area = alpha * sm.unit_area[sm.unit_of(p, s)]
                    ws.cell(row, 2 + j).value = round(area, 3) if area > 0 else None
    wb.save(out_path)
    print(f'已保存: {out_path}')


if __name__ == '__main__':
    import pickle

    sm, base = build_baseline()

    # 问题1 两情形：加载全量权重 DE 求解器保存的最终方案（见 src/de_solver_full.py __main__）
    scenarios = []
    for mode in (1, 2):
        with open(f'{OUT}/de_plan_full_mode{mode}.pkl', 'rb') as f:
            plan = pickle.load(f)
        # 检查优先：fitness 先过规则检查器（2023 固定基线），违规即 raise
        fit = FullDE(seed=0).fitness(plan, problem=1, mode=mode, base=base)
        scenarios.append((plan, f'result1_{mode}.xlsx'))
        print(f'问题1 mode={mode}（全量权重 DE 方案）: {fit / 1e4:.2f} 万\n')

    de = FullDE(seed=0)
    plan, fit = de.solve(baseline=base, problem=2, mode=1, dist='normal',
                         seed=0, n_quad=16)
    de.report(plan, mode=1)
    scenarios.append((plan, 'result2.xlsx'))
    print(f'问题2 normal: {fit / 1e4:.2f} 万\n')

    for plan, name in scenarios:
        fill_workbook(f'{DOCS}/{name}', f'{OUT}/{name}', plan)

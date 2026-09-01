# -*- coding: utf-8 -*-
"""
板凳龙模拟器（2024 高教社杯 A 题）
==================================================================
单一模拟器模块，包含：
  1. 模拟核心    —— 阿基米德螺线 + 刚性板凳链（真实板凳几何：长/宽/锚点）
  2. 可选参数    —— 螺距、龙头速度、初始圈数、板凳规格等（SimParams）
  3. 可视化      —— matplotlib 静态图 / GIF / 状态导出
  4. 求解器接口  —— Solver 基类 + 注册机制，预留给问题 1~4 求解器接入

板凳真实几何
------------
每节板凳为矩形木板，两端各钻一孔（锚点），孔心距板端 27.5cm，孔径 5.5cm。
相邻板凳在孔处用销铰接，把手即孔（锚点）位置。

  龙头板    : 长 341cm，宽 30cm，孔心距 = 341-2*27.5 = 286cm = 2.86m
  龙身/龙尾 : 长 220cm，宽 30cm，孔心距 = 220-2*27.5 = 165cm = 1.65m

把手编号（共 224 个）
---------------------
h0    = 龙头前把手（龙头板前孔）
hk    = 第 k 节龙身前把手（k=1..221）
h222  = 龙尾前把手
h223  = 龙尾后把手
链段 k 为把手 k -> 把手 k+1（k=0 为龙头板孔心距 2.86m，其余 1.65m）。

模型假设
--------
1. 盘入螺线为等距（阿基米德）螺线：r(theta) = a*theta，a = 螺距/(2*pi)。
2. 龙头前把手初始位于螺线第 16 圈 A 点：theta0 = 32*pi，r0 = 8.8m，(8.8, 0)。
3. 龙头前把手以 1 m/s 顺时针恒速盘入（theta 随时间减小）。
4. 各把手中心位于螺线上，同节板凳两孔（锚点）中心距固定（刚性）。
5. 速度由锚点约束隐式求导解析获得。
"""
import os
import math
import numpy as np

BASE = os.path.dirname(os.path.abspath(__file__))

# ====================================================================
# 一、板凳几何规格
# ====================================================================
class BenchSpec:
    """一节板凳的几何规格：板长、板宽、锚点（孔）距板端的距离。"""

    __slots__ = ("length", "width", "hole_inset")

    def __init__(self, length, width, hole_inset):
        self.length = length      # 板长 (m)
        self.width = width        # 板宽 (m)
        self.hole_inset = hole_inset  # 孔心距板端 (m)

    @property
    def hole_distance(self):
        """两锚点（孔心）之间的距离。"""
        return self.length - 2.0 * self.hole_inset

    def __repr__(self):
        return (f"BenchSpec(len={self.length}, w={self.width}, "
                f"inset={self.hole_inset}, hole_d={self.hole_distance})")


# 题目标准板凳规格
HEAD_BENCH = BenchSpec(3.41, 0.30, 0.275)   # 龙头板
BODY_BENCH = BenchSpec(2.20, 0.30, 0.275)   # 龙身 / 龙尾板

# ====================================================================
# 二、模拟可选参数
# ====================================================================
class SimParams:
    """模拟可选参数。所有字段均有默认值，可仅覆盖需要的项。"""

    __slots__ = ("pitch", "v_head", "turns", "n_head", "n_body",
                 "n_tail", "t_max", "head_bench", "body_bench", "collision_gap")

    def __init__(self, pitch=0.55, v_head=1.0, turns=16,
                 n_head=1, n_body=221, n_tail=1, t_max=600.0,
                 head_bench=HEAD_BENCH, body_bench=BODY_BENCH,
                 collision_gap=1e-6):
        self.pitch = pitch            # 螺距 (m)
        self.v_head = v_head          # 龙头前把手线速度 (m/s)
        self.turns = turns            # 龙头前把手初始圈数
        self.n_head = n_head          # 龙头板数量
        self.n_body = n_body          # 龙身板数量
        self.n_tail = n_tail          # 龙尾板数量
        self.t_max = t_max            # 最大模拟时间 (s)
        self.head_bench = head_bench  # 龙头板规格
        self.body_bench = body_bench  # 龙身/龙尾板规格
        self.collision_gap = collision_gap  # 碰撞判据：非相邻板间距 (m)

    @property
    def n_boards(self):
        """总板数。"""
        return self.n_head + self.n_body + self.n_tail

    @property
    def n_handles(self):
        """总把手（锚点）数 = 总板数 + 1。"""
        return self.n_boards + 1

    @property
    def theta0(self):
        """龙头前把手初始极角。"""
        return self.turns * 2.0 * np.pi

    def bench_of(self, i):
        """第 i 块板（i=0..n_boards-1）的规格。"""
        return self.head_bench if i < self.n_head else self.body_bench

    def link_of(self, k):
        """链段 k（把手 k -> 把手 k+1）的孔心距。"""
        return self.bench_of(k).hole_distance

    def __repr__(self):
        return (f"SimParams(pitch={self.pitch}, v_head={self.v_head}, "
                f"turns={self.turns}, n_handles={self.n_handles})")


# ====================================================================
# 三、螺线几何（纯标量实现，避免 numpy 标量开销）
# ====================================================================
def make_spiral(pitch):
    """螺线参数 a：r = a * theta。"""
    return pitch / (2.0 * math.pi)


def pos(a, theta):
    """螺线上极角 theta 处的直角坐标。"""
    r = a * theta
    return (r * math.cos(theta), r * math.sin(theta))


def dpos_dtheta(a, theta):
    """极角导数 dP/dtheta。"""
    r = a * theta
    return (a * math.cos(theta) - r * math.sin(theta),
            a * math.sin(theta) + r * math.cos(theta))


def arc(a, theta):
    """螺线从 theta=0 到 theta 的弧长。"""
    t = float(theta)
    return a / 2.0 * (t * math.sqrt(t * t + 1.0) + math.asinh(t))


def theta_from_arc(a, s, tol=1e-12):
    """弧长反解极角（牛顿迭代，初值 s/a）。"""
    t = s / a
    for _ in range(100):
        f = arc(a, t) - s
        if abs(f) < tol:
            break
        t -= f / (a * math.sqrt(t * t + 1.0))   # d(arc)/dt = a*sqrt(t^2+1)
    return t


def next_theta(a, prev, link, guess):
    """
    已知把手极角 prev，求下一把手极角（theta>prev）使弦长 = link。
    牛顿法；guess 为上一时刻同把手极角（热启动）。
    f(d) = |P(prev+d) - P(prev)| - link
    """
    d = guess - prev
    if d <= 0:
        d = link / (a * prev)                 # 初始估计：弦 ≈ r*dθ
    for _ in range(30):
        p0 = pos(a, prev)
        p1 = pos(a, prev + d)
        dx = p1[0] - p0[0]
        dy = p1[1] - p0[1]
        n = math.hypot(dx, dy)
        f = n - link
        if abs(f) < 1e-12:
            break
        # 导数：df/dd = (p1-p0)·P'(prev+d) / |p1-p0|
        dp = dpos_dtheta(a, prev + d)
        df = (dx * dp[0] + dy * dp[1]) / n
        if abs(df) < 1e-15:
            break
        d -= f / df
        if d <= 0:
            d = link / (a * prev)
    return prev + d


# ====================================================================
# 四、模拟状态（一次求解的结果）
# ====================================================================
class SimState:
    """某个时刻的完整状态。

    属性
    ----
    t       : 时刻 (s)
    theta0  : 龙头前把手极角
    thetas  : 全部把手极角 (n_handles,)
    handles : 全部把手（锚点）直角坐标 (n_handles, 2)
    speeds  : 全部把手线速度 (n_handles,)
    boards  : 各板矩形四顶点（真实长宽含伸出段） (n_boards, 4, 2)
    anchors : 各板两锚点（孔）坐标 (n_boards, 2, 2)
    dmin    : 非相邻板间最小间距 (m)
    collided: 是否发生碰撞
    """

    def __init__(self, t, theta0, thetas, handles, speeds, boards, anchors,
                 dmin=float("inf"), collided=False):
        self.t = t
        self.theta0 = theta0
        self.thetas = thetas
        self.handles = handles
        self.speeds = speeds
        self.boards = boards
        self.anchors = anchors
        self.dmin = dmin
        self.collided = collided

    def copy(self, **over):
        d = dict(t=self.t, theta0=self.theta0, thetas=self.thetas,
                 handles=self.handles, speeds=self.speeds,
                 boards=self.boards, anchors=self.anchors,
                 dmin=self.dmin, collided=self.collided)
        d.update(over)
        return SimState(**d)


# ====================================================================
# 五、模拟器
# ====================================================================
class Simulator:
    """板凳龙盘入模拟器。

    用法
    ----
    sim = Simulator(SimParams(pitch=0.55, v_head=1.0, turns=16))
    st = sim.state(t=60)            # 某时刻状态
    Thetas, Pos, Speed = sim.trajectory(times)   # 批量轨迹
    """

    def __init__(self, params=None):
        self.params = params or SimParams()
        p = self.params
        self.a = make_spiral(p.pitch)
        self.s0 = arc(self.a, p.theta0)          # 龙头前把手初始弧长
        self.links = [p.link_of(k) for k in range(p.n_boards)]
        self.nh = p.n_handles
        # 连续播放热启动：上一时刻的把手极角
        self._prev_row = None

    # ---- 单时刻状态 ----
    def state(self, t=0.0, use_warm=True):
        p = self.params
        s = max(self.s0 - p.v_head * t, 0.0)
        th0 = theta_from_arc(self.a, s)
        warm = self._prev_row if (use_warm and self._prev_row is not None) else None
        thetas = np.empty(self.nh)
        thetas[0] = th0
        for k in range(1, self.nh):
            guess = warm[k] if warm is not None else thetas[k - 1]
            thetas[k] = next_theta(self.a, thetas[k - 1], self.links[k - 1], guess)
        self._prev_row = thetas

        handles = np.stack([pos(self.a, th) for th in thetas])
        dth0 = -p.v_head / np.linalg.norm(dpos_dtheta(self.a, th0))
        speeds = self._speeds(thetas, dth0)
        boards, anchors = self._boards(handles)
        dmin, collided = self._min_gap(boards, thetas)
        return SimState(t, th0, thetas, handles, speeds, boards, anchors,
                        dmin=dmin, collided=collided)

    def reset(self):
        """清除热启动缓存（参数或时刻跳变后调用）。"""
        self._prev_row = None

    # ---- 速度：锚点约束隐式求导 ----
    def _speeds(self, thetas, dtheta0):
        P = np.stack([pos(self.a, th) for th in thetas])
        dP = np.stack([dpos_dtheta(self.a, th) for th in thetas])
        w = np.empty(self.nh)
        w[0] = dtheta0
        for k in range(1, self.nh):
            vdiff = P[k] - P[k - 1]
            w[k] = w[k - 1] * (vdiff @ dP[k - 1]) / (vdiff @ dP[k])
        speeds = np.sqrt(np.sum(dP ** 2, axis=1)) * np.abs(w)
        return speeds

    # ---- 真实板凳几何：矩形 + 锚点 ----
    def _boards(self, handles):
        """
        由把手（锚点）位置计算每块板：
        板中心 = 两锚点中点；板沿锚点连线方向向两端各延伸 hole_inset；
        板宽方向 ± width/2。返回 (矩形顶点 (n,4,2), 锚点 (n,2,2))。
        """
        p = self.params
        nb = p.n_boards
        verts = np.empty((nb, 4, 2))
        anchors = np.empty((nb, 2, 2))
        for i in range(nb):
            spec = p.bench_of(i)
            p1, p2 = handles[i], handles[i + 1]
            dx, dy = p2 - p1
            ln = np.hypot(dx, dy)
            ux, uy = dx / ln, dy / ln        # 板长方向单位向量
            nx, ny = -uy, ux                 # 板宽方向单位向量
            cx, cy = (p1[0] + p2[0]) / 2, (p1[1] + p2[1]) / 2
            hl, hw = spec.length / 2.0, spec.width / 2.0
            verts[i, 0] = [cx + hl * ux + hw * nx, cy + hl * uy + hw * ny]
            verts[i, 1] = [cx + hl * ux - hw * nx, cy + hl * uy - hw * ny]
            verts[i, 2] = [cx - hl * ux - hw * nx, cy - hl * uy - hw * ny]
            verts[i, 3] = [cx - hl * ux + hw * nx, cy - hl * uy + hw * ny]
            anchors[i, 0] = p1
            anchors[i, 1] = p2
        return verts, anchors

    # ---- 碰撞：非相邻板矩形最小距离（接触判据，全向量化） ----
    def _min_gap(self, boards, thetas=None):
        """
        非相邻板（索引差 >= 2）的最小距离。
        碰撞判据：两板矩形接触/干涉（距离 < collision_gap）。
        注意：相邻圈板间径向间隙 ≈ 螺距 - 板宽 = 0.25m，因此不能用
        "间距 < 板宽" 判碰撞，必须用真实矩形接触判据。
        thetas: 把手极角，用于极角对齐粗筛（同径向方向的相邻圈板才可能接近）。
        """
        gap = self.params.collision_gap
        nb = boards.shape[0]
        if nb < 3:
            return float("inf"), False
        cxs = boards.mean(axis=1)[:, 0]      # 板中心（粗筛用）
        cys = boards.mean(axis=1)[:, 1]
        idx = np.arange(nb)
        # 候选板对：极角对齐 + 中心距粗筛（向量化）
        ok = np.hypot(cxs[:, None] - cxs[None, :],
                      cys[:, None] - cys[None, :]) < 3.7
        if thetas is not None:
            cth = (thetas[:-1] + thetas[1:]) / 2.0
            dth = np.abs(cth[:, None] - cth[None, :])
            wrap = dth % (2.0 * np.pi)
            ok &= np.minimum(wrap, 2.0 * np.pi - wrap) <= 1.2
        ok &= np.abs(idx[:, None] - idx[None, :]) >= 2   # 非相邻板
        ii, jj = np.where(np.triu(ok, k=2))
        if ii.size == 0:
            return float("inf"), False
        d = Simulator._poly_dist_vec(boards[ii], boards[jj])
        dmin = float(d.min())
        return dmin, (dmin < gap)

    @staticmethod
    def _poly_dist_vec(A, B):
        """
        A, B: (n, 4, 2) 凸四边形顶点数组，向量化计算两两最小距离。
        相交（分离轴检验）时为 0；否则取 8 种"点-边"距离的最小值。
        """
        n = A.shape[0]
        # ---- 分离轴检验（凸多边形相交判定）----
        Ae = np.roll(A, -1, axis=1) - A            # (n,4,2) 边向量
        Be = np.roll(B, -1, axis=1) - B
        axes = np.concatenate([Ae[:, :2], Be[:, :2]], axis=1)   # 两矩形各取 2 个独立边方向
        axes = np.stack([-axes[..., 1], axes[..., 0]], axis=-1)  # 边法线
        axes = axes / np.maximum(np.hypot(axes[..., 0], axes[..., 1])[..., None], 1e-30)
        pA = (A[:, None, :, :] * axes[:, :, None, :]).sum(-1)   # (n,4,4) [对, 轴, 顶点]
        pB = (B[:, None, :, :] * axes[:, :, None, :]).sum(-1)
        sep = (pA.max(-1) < pB.min(-1)) | (pB.max(-1) < pA.min(-1))
        inter = ~sep.any(axis=1)                   # 任一轴不分离 -> 相交

        # ---- 点-边最小距离（4 顶点 x 4 边，双向）----
        def pt_edges(P, V, E):
            # P (n,4,2) 顶点, V (n,4,2) 边起点(即多边形顶点), E (n,4,2) 边向量
            s1 = V[:, None, :, :]              # (n,1,4,2) 边起点
            d = E[:, None, :, :]               # (n,1,4,2) 边向量
            L2 = (d * d).sum(-1)               # (n,4,1)
            Pc = P[:, :, None, :]              # (n,4,1,2)
            tt = np.clip(((Pc - s1) * d).sum(-1) / np.maximum(L2, 1e-30), 0.0, 1.0)
            proj = s1 + tt[..., None] * d      # (n,4,4,2)
            dist = np.hypot((Pc - proj)[..., 0], (Pc - proj)[..., 1])
            return dist.min(axis=(1, 2))       # (n,)

        d = np.minimum(pt_edges(A, B, Be), pt_edges(B, A, Ae))
        return np.where(inter, 0.0, d)

    # ---- 批量轨迹 ----
    def trajectory(self, times, reset=True):
        """批量模拟：返回 (极角 (T,n), 位置 (T,n,2), 速度 (T,n))。"""
        if reset:
            self.reset()
        T = len(times)
        Thetas = np.empty((T, self.nh))
        Pos = np.empty((T, self.nh, 2))
        Speed = np.empty((T, self.nh))
        for i, t in enumerate(times):
            st = self.state(t)
            Thetas[i] = st.thetas
            Pos[i] = st.handles
            Speed[i] = st.speeds
        return Thetas, Pos, Speed


# ====================================================================
# 六、可视化（matplotlib，可选依赖）
# ====================================================================
def _setup_chinese_font():
    """配置 matplotlib 中文字体（Windows 常见字体）。"""
    import matplotlib
    candidates = ["Microsoft YaHei", "SimHei", "SimSun", "PingFang SC",
                  "Noto Sans CJK SC", "WenQuanYi Micro Hei"]
    available = {f.name for f in matplotlib.font_manager.fontManager.ttflist}
    for name in candidates:
        if name in available:
            matplotlib.rcParams["font.sans-serif"] = [name]
            break
    matplotlib.rcParams["font.sans-serif"] += ["DejaVu Sans"]
    matplotlib.rcParams["axes.unicode_minus"] = False


def plot_state(sim, t=0.0, ax=None, show_spiral=True, show_anchors=True,
               title=None, figsize=(8, 8)):
    """绘制某时刻的板凳龙（真实长宽矩形 + 锚点 + 螺线）。

    返回 matplotlib 的 Figure。
    """
    import matplotlib.pyplot as plt
    _setup_chinese_font()
    st = sim.state(t)
    p = sim.params
    a = sim.a

    if ax is None:
        fig, ax = plt.subplots(figsize=figsize)
    else:
        fig = ax.figure
    ax.set_aspect("equal")

    # 螺线
    if show_spiral:
        tmax = theta_from_arc(a, max(sim.s0 - p.v_head * t, 0)) + \
            p.turns * 2 * np.pi + 6 * np.pi
        th = np.linspace(0, tmax, 4000)
        R = a * th
        ax.plot(R * np.cos(th), R * np.sin(th), color="#7896be",
                lw=0.8, alpha=0.4, zorder=1)

    # 板凳矩形（真实长宽）
    colors = []
    for i in range(p.n_boards):
        if i < p.n_head:
            c = "#ff7a45"
        elif i == p.n_boards - 1:
            c = "#5ee08a"
        else:
            c = "#4da3ff"
        colors.append(c)
    for i in range(p.n_boards):
        poly = st.boards[i]
        px = np.append(poly[:, 0], poly[0, 0])
        py = np.append(poly[:, 1], poly[0, 1])
        ax.fill(px, py, color=colors[i], alpha=0.55,
                edgecolor=colors[i], lw=0.8, zorder=2)

    # 锚点（孔）
    if show_anchors:
        ax.plot(st.handles[:, 0], st.handles[:, 1], "o", ms=2.4,
                color="#fff", mec="#333", mew=0.4, zorder=3)

    # 龙头 / 龙尾标记
    ax.plot(*st.handles[0], "o", ms=9, color="#ff4d4d",
            mec="#fff", mew=1.5, zorder=4)
    ax.plot(*st.handles[-1], "o", ms=8, color="#5ee08a",
            mec="#fff", mew=1.5, zorder=4)

    ax.set_title(title or f"t = {t:.1f} s   龙头半径 r0 = "
                 f"{np.hypot(*st.handles[0]):.3f} m")
    ax.set_xlabel("x (m)")
    ax.set_ylabel("y (m)")
    ax.grid(alpha=0.25)
    ax.plot([0], [0], "+", color="#888", ms=8, zorder=5)
    return fig


def animate(sim, times=None, out_path=None, fps=8, dpi=110, follow=False):
    """生成盘入过程 GIF（真实板凳矩形）。返回保存路径。"""
    import matplotlib.pyplot as plt
    from matplotlib.animation import FuncAnimation
    _setup_chinese_font()
    p = sim.params
    if times is None:
        times = np.arange(0, 301, 5.0)
    if out_path is None:
        out_path = os.path.join(BASE, "dragon.gif")

    fig, ax = plt.subplots(figsize=(8, 8))
    Rmax = a_max = p.theta0 * sim.a
    half = max(Rmax, 3) + 2.5

    def draw(t):
        ax.clear()
        st = sim.state(t)
        # 视口
        if follow:
            r0 = np.hypot(*st.handles[0])
            hl = max(4.0, r0 + 7)
            cx, cy = st.handles[0]
            ax.set_xlim(cx - hl, cx + hl)
            ax.set_ylim(cy - hl, cy + hl)
        else:
            ax.set_xlim(-half, half)
            ax.set_ylim(-half, half)
        ax.set_aspect("equal")
        # 螺线
        tmax = theta_from_arc(sim.a, max(sim.s0 - p.v_head * t, 0)) + \
            p.turns * 2 * np.pi + 6 * np.pi
        th = np.linspace(0, tmax, 3000)
        R = sim.a * th
        ax.plot(R * np.cos(th), R * np.sin(th), color="#7896be",
                lw=0.7, alpha=0.35, zorder=1)
        # 板凳
        for i in range(p.n_boards):
            poly = st.boards[i]
            px = np.append(poly[:, 0], poly[0, 0])
            py = np.append(poly[:, 1], poly[0, 1])
            c = ("#ff7a45" if i == 0 else
                 "#5ee08a" if i == p.n_boards - 1 else "#4da3ff")
            ax.fill(px, py, color=c, alpha=0.55, edgecolor=c, lw=0.7, zorder=2)
        ax.plot(st.handles[:, 0], st.handles[:, 1], "o", ms=2,
                color="#fff", mec="#333", mew=0.3, zorder=3)
        ax.plot(*st.handles[0], "o", ms=9, color="#ff4d4d",
                mec="#fff", mew=1.5, zorder=4)
        ax.plot(*st.handles[-1], "o", ms=8, color="#5ee08a",
                mec="#fff", mew=1.5, zorder=4)
        ax.set_title(f"t = {t:.1f} s   龙头半径 r0 = "
                     f"{np.hypot(*st.handles[0]):.3f} m")
        ax.grid(alpha=0.2)

    anim = FuncAnimation(fig, draw, frames=times, interval=1000 / fps)
    anim.save(out_path, writer="pillow", fps=fps, dpi=dpi)
    plt.close(fig)
    print(f"[已保存] {out_path}")
    return out_path


# ====================================================================
# 七、结果导出（附件模板格式）
# ====================================================================
HANDLE_ROW_LABELS = (["龙头"] + [f"第{i}节龙身" for i in range(1, 222)]
                     + ["龙尾", "龙尾（后）"])


def write_result1(Pos, Speed, times, out_path):
    """问题1：位置表 (448 行 x 301 列) + 速度表 (224 行 x 301 列)。"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "位置"
    ws.append(["时间"] + [f"{int(t)} s" for t in times])
    labels = [f"{h}x (m)" for h in HANDLE_ROW_LABELS] + \
             [f"{h}y (m)" for h in HANDLE_ROW_LABELS]
    for lab, row in zip(labels, Pos.reshape(-1, Pos.shape[1])):
        ws.append([lab] + [f"{v:.6f}" for v in row])
    ws2 = wb.create_sheet("速度")
    ws2.append(["时间"] + [f"{int(t)} s" for t in times])
    for lab, row in zip(HANDLE_ROW_LABELS, Speed):
        ws2.append([f"{lab} (m/s)"] + [f"{v:.6f}" for v in row])
    wb.save(out_path)
    print(f"[已保存] {out_path}")


def write_result2(Pos, Speed, out_path):
    """问题2：终止时刻，224 行 x [把手, x, y, 速度]。"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["把手", "横坐标x (m)", "纵坐标y (m)", "速度 (m/s)"])
    for lab, p_, v in zip(HANDLE_ROW_LABELS, Pos, Speed):
        ws.append([lab, f"{p_[0]:.6f}", f"{p_[1]:.6f}", f"{v:.6f}"])
    wb.save(out_path)
    print(f"[已保存] {out_path}")


# ====================================================================
# 八、求解器接口（预留）
# ====================================================================
class Solver:
    """
    求解器接口（预留）。

    问题 1~4 的求解器通过继承 Solver 并实现 solve() 接入模拟器，
    然后用 register_solver 注册，便于统一调用：

        @register_solver
        class ProblemXSolver(Solver):
            name = "problemX"
            def solve(self, sim, **kwargs):
                ...   # 读取 sim.params / sim.state(t) 等
                return result

    内置示例：Problem1Solver（位置/速度表格）、Problem2Solver（碰撞时刻）。
    """

    name = "base"

    def solve(self, sim, **kwargs):
        raise NotImplementedError


_SOLVERS = {}


def register_solver(cls):
    """装饰器：注册求解器类，键为 cls.name。"""
    _SOLVERS[cls.name] = cls
    return cls


def get_solver(name):
    """按名称获取求解器类。"""
    return _SOLVERS.get(name)


def list_solvers():
    """已注册的求解器名称列表。"""
    return list(_SOLVERS)


@register_solver
class Problem1Solver(Solver):
    """问题1：龙头前把手 1 m/s 盘入 0~300s 的把手位置/速度，输出附件模板。"""

    name = "problem1"

    def solve(self, sim, times=None, out_xlsx=None):
        if times is None:
            times = np.arange(0, 301, dtype=float)
        if out_xlsx is None:
            out_xlsx = os.path.join(BASE, "result1.xlsx")
        _, Pos, Speed = sim.trajectory(times)
        write_result1(Pos, Speed, times, out_xlsx)
        return Pos, Speed


@register_solver
class Problem2Solver(Solver):
    """
    问题2：二分求首次碰撞终止时刻。
    判据：非相邻板（真实矩形，含板头伸出段）接触/干涉（距离 < 1e-6）。
    注：深盘入段存在几何分支跳变，t 上限取 460s。
    """

    name = "problem2"

    def solve(self, sim, t_max=460.0, out_xlsx=None):
        p = sim.params
        gap = p.collision_gap

        # 注意：必须每次全新求解（reset 清热启动）。热启动会使链条解
        # 锁死在错误分支，深盘入段无法检测到真实碰撞。
        def state(t):
            sim.reset()
            return sim.state(t)

        # 两阶段粗扫：5s -> 1s 步长找到含碰撞的 1s 区间，再二分
        def first_collided(lo, hi, step):
            t = lo
            while t <= hi:
                if state(t).dmin < gap:
                    return t
                t += step
            return None

        t1 = first_collided(0.0, t_max, 5.0)
        if t1 is None:
            raise RuntimeError(f"在 {t_max}s 内未检测到碰撞")
        t2 = first_collided(t1 - 5.0, t1, 1.0)
        t_lo, t_hi = t2 - 1.0, t2
        for _ in range(30):
            tm = 0.5 * (t_lo + t_hi)
            if state(tm).dmin < gap:
                t_hi = tm
            else:
                t_lo = tm
        tc = 0.5 * (t_lo + t_hi)
        st = state(tc)
        if out_xlsx is None:
            out_xlsx = os.path.join(BASE, "result2.xlsx")
        write_result2(st.handles, st.speeds, out_xlsx)
        return tc, st


# ====================================================================
# 九、演示 / 自测
# ====================================================================
if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description="板凳龙模拟器")
    ap.add_argument("--pitch", type=float, default=0.55, help="螺距 (m)")
    ap.add_argument("--v", type=float, default=1.0, help="龙头速度 (m/s)")
    ap.add_argument("--turns", type=int, default=16, help="初始圈数")
    ap.add_argument("--t", type=float, default=0.0, help="绘制时刻 (s)")
    ap.add_argument("--plot", action="store_true", help="绘制静态图")
    ap.add_argument("--animate", action="store_true", help="生成 GIF")
    ap.add_argument("--list-solvers", action="store_true", help="列出求解器")
    args = ap.parse_args()

    print("== 板凳龙模拟器 ==")
    print("已注册求解器:", list_solvers())

    params = SimParams(pitch=args.pitch, v_head=args.v, turns=args.turns)
    sim = Simulator(params)
    print(f"参数: {params}")
    print(f"  a = {sim.a:.6f}, s0 = {sim.s0:.3f} m")

    st = sim.state(args.t)
    r0 = np.hypot(*st.handles[0])
    print(f"t = {args.t:.1f} s: 龙头极角 {st.theta0:.4f} rad, "
          f"半径 {r0:.4f} m, 速度 {st.speeds[0]:.6f} m/s, "
          f"尾把手速度 {st.speeds[-1]:.6f} m/s, "
          f"最小非相邻板间距 {st.dmin:.4f} m")

    if args.list_solvers:
        import sys
        sys.exit(0)

    if args.plot:
        plot_state(sim, t=args.t, show_spiral=True, show_anchors=True)
        import matplotlib.pyplot as plt
        plt.show()

    if args.animate:
        animate(sim, times=np.arange(0, 301, 5.0),
                out_path=os.path.join(BASE, "dragon.gif"))
